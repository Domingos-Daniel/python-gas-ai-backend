"""
Utilitários para exportação de dados em formatos CSV e Excel.
Fornece funcionalidades para exportar dados de análises, gráficos e conversas.
"""

import csv
import io
import json
import logging
from typing import Dict, Any, List, Optional, Union
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logger = logging.getLogger(__name__)


class DataExporter:
    """Classe para exportação de dados em diferentes formatos."""
    
    def __init__(self):
        """Inicializa o exportador de dados."""
        self.supported_formats = ['csv', 'xlsx', 'json']
    
    def export_chat_history(self, messages: List[Dict[str, Any]], format_type: str = 'xlsx') -> bytes:
        """
        Exporta o histórico de conversas do chat.
        
        Args:
            messages: Lista de mensagens do chat
            format_type: Formato de exportação ('csv', 'xlsx', 'json')
            
        Returns:
            Bytes do arquivo exportado
        """
        try:
            # Prepara dados para exportação
            export_data = []
            for msg in messages:
                export_data.append({
                    'Data/Hora': msg.get('timestamp', datetime.now()).strftime('%Y-%m-%d %H:%M:%S') if hasattr(msg.get('timestamp'), 'strftime') else str(msg.get('timestamp', '')),
                    'Tipo': msg.get('role', 'unknown'),
                    'Conteúdo': msg.get('content', ''),
                    'Tem Gráficos': msg.get('hasCharts', False),
                    'ID': msg.get('id', '')
                })
            
            if format_type == 'csv':
                return self._export_to_csv(export_data, 'Chat_Historico')
            elif format_type == 'xlsx':
                return self._export_to_excel(export_data, 'Chat_Historico')
            elif format_type == 'json':
                return self._export_to_json(messages)
            else:
                raise ValueError(f"Formato não suportado: {format_type}")
                
        except Exception as e:
            logger.error(f"Erro ao exportar histórico de chat: {e}")
            raise
    
    def export_analysis_data(self, analysis_result: Dict[str, Any], format_type: str = 'xlsx') -> bytes:
        """
        Exporta dados de análise com gráficos e métricas.
        
        Args:
            analysis_result: Resultado da análise com dados e gráficos
            format_type: Formato de exportação ('csv', 'xlsx', 'json')
            
        Returns:
            Bytes do arquivo exportado
        """
        try:
            # Extrai dados da análise
            export_sheets = {}
            
            # Dados principais da análise
            main_data = {
                'Título': [analysis_result.get('title', 'Análise de Dados')],
                'Subtítulo': [analysis_result.get('subtitle', '')],
                'Categoria': [analysis_result.get('analysis_category', 'general')],
                'Confiança': [analysis_result.get('confidence', 0.8)],
                'Total de Itens': [analysis_result.get('total_items', 0)],
                'Data da Análise': [datetime.now().strftime('%Y-%m-%d %H:%M:%S')]
            }
            export_sheets['Informações_Gerais'] = pd.DataFrame(main_data)
            
            # Dados de contexto/análise
            contextual_data = analysis_result.get('contextual_analysis', {})
            if contextual_data:
                context_df = pd.DataFrame([{
                    'Resumo Executivo': contextual_data.get('executive_summary', ''),
                    'Título': contextual_data.get('title', ''),
                    'Confiança': contextual_data.get('confidence', 0.8)
                }])
                export_sheets['Análise_Contextual'] = context_df
            
            # KPIs
            kpis = analysis_result.get('kpis', {})
            if kpis:
                kpi_data = []
                for kpi_name, kpi_info in kpis.items():
                    kpi_data.append({
                        'KPI': kpi_name.replace('_', ' ').title(),
                        'Valor Atual': kpi_info.get('current', 0),
                        'Valor Alvo': kpi_info.get('target', 0),
                        'Status': kpi_info.get('status', 'unknown'),
                        'Tendência': kpi_info.get('trend', 'stable')
                    })
                export_sheets['KPIs'] = pd.DataFrame(kpi_data)
            
            # Tendências
            trends = analysis_result.get('trends', {})
            if trends:
                trend_data = []
                for period, trend_list in trends.items():
                    if isinstance(trend_list, list):
                        for trend in trend_list:
                            trend_data.append({
                                'Período': period,
                                'Tendência': trend
                            })
                if trend_data:
                    export_sheets['Tendências'] = pd.DataFrame(trend_data)
            
            # Recomendações
            recommendations = analysis_result.get('recommendations', [])
            if recommendations:
                rec_data = []
                for rec in recommendations:
                    rec_data.append({
                        'Categoria': rec.get('category', 'Geral'),
                        'Prioridade': rec.get('priority', 'Média'),
                        'Recomendação': rec.get('recommendation', ''),
                        'Impacto': rec.get('impact', 'Não especificado')
                    })
                export_sheets['Recomendações'] = pd.DataFrame(rec_data)
            
            # Dados principais (gráficos)
            main_chart_data = analysis_result.get('data', {})
            if main_chart_data and isinstance(main_chart_data, dict):
                chart_df = pd.DataFrame(list(main_chart_data.items()), columns=['Categoria', 'Valor'])
                export_sheets['Dados_Principais'] = chart_df
            
            if format_type == 'csv':
                # Para CSV, combina todas as sheets em um único arquivo
                combined_data = []
                for sheet_name, df in export_sheets.items():
                    df_copy = df.copy()
                    df_copy['Sheet'] = sheet_name
                    combined_data.append(df_copy)
                
                if combined_data:
                    final_df = pd.concat(combined_data, ignore_index=True)
                    return self._export_dataframe_to_csv(final_df, 'Análise_Completa')
                else:
                    return self._export_dataframe_to_csv(pd.DataFrame(), 'Análise_Vazia')
                    
            elif format_type == 'xlsx':
                return self._export_multiple_sheets_to_excel(export_sheets, 'Análise_Completa')
            elif format_type == 'json':
                return self._export_to_json(analysis_result)
            else:
                raise ValueError(f"Formato não suportado: {format_type}")
                
        except Exception as e:
            logger.error(f"Erro ao exportar dados de análise: {e}")
            raise
    
    def export_chart_data(self, chart_data: Dict[str, Any], format_type: str = 'xlsx') -> bytes:
        """
        Exporta dados de gráficos individuais.
        
        Args:
            chart_data: Dados do gráfico com labels e valores
            format_type: Formato de exportação ('csv', 'xlsx', 'json')
            
        Returns:
            Bytes do arquivo exportado
        """
        try:
            # Prepara dados do gráfico
            labels = chart_data.get('labels', [])
            values = chart_data.get('values', [])
            chart_type = chart_data.get('chart_type', 'unknown')
            title = chart_data.get('title', 'Gráfico de Dados')
            
            export_data = []
            for i, (label, value) in enumerate(zip(labels, values)):
                export_data.append({
                    'Categoria': label,
                    'Valor': value,
                    'Tipo de Gráfico': chart_type,
                    'Título': title,
                    'Índice': i + 1
                })
            
            if format_type == 'csv':
                return self._export_to_csv(export_data, f'Gráfico_{chart_type}')
            elif format_type == 'xlsx':
                return self._export_to_excel(export_data, f'Gráfico_{chart_type}')
            elif format_type == 'json':
                return self._export_to_json(chart_data)
            else:
                raise ValueError(f"Formato não suportado: {format_type}")
                
        except Exception as e:
            logger.error(f"Erro ao exportar dados de gráfico: {e}")
            raise
    
    def _export_to_csv(self, data: List[Dict[str, Any]], filename_prefix: str) -> bytes:
        """Exporta dados para formato CSV."""
        output = io.StringIO()
        if data:
            writer = csv.DictWriter(output, fieldnames=data[0].keys())
            writer.writeheader()
            writer.writerows(data)
        
        csv_content = output.getvalue()
        output.close()
        return csv_content.encode('utf-8')
    
    def _export_dataframe_to_csv(self, df: pd.DataFrame, filename_prefix: str) -> bytes:
        """Exporta DataFrame para formato CSV."""
        output = io.StringIO()
        df.to_csv(output, index=False, encoding='utf-8')
        csv_content = output.getvalue()
        output.close()
        return csv_content.encode('utf-8')
    
    def _export_to_excel(self, data: List[Dict[str, Any]], sheet_name: str) -> bytes:
        """Exporta dados para formato Excel com uma única aba."""
        if not data:
            data = [{'Mensagem': 'Nenhum dado disponível'}]
        
        df = pd.DataFrame(data)
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name[:31], index=False)  # Limite de 31 caracteres para nomes de abas
            
            # Formatação básica
            workbook = writer.book
            worksheet = writer.sheets[sheet_name[:31]]
            
            # Formata cabeçalho
            for cell in worksheet[1]:
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            # Ajusta largura das colunas
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)  # Máximo de 50 caracteres
                worksheet.column_dimensions[column_letter].width = adjusted_width
        
        excel_content = output.getvalue()
        output.close()
        return excel_content
    
    def _export_multiple_sheets_to_excel(self, sheets_data: Dict[str, pd.DataFrame], filename_prefix: str) -> bytes:
        """Exporta múltiplas abas para Excel."""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in sheets_data.items():
                if df.empty:
                    df = pd.DataFrame({'Mensagem': ['Nenhum dado disponível para esta seção']})
                
                # Limita nome da aba para 31 caracteres
                excel_sheet_name = sheet_name[:31]
                df.to_excel(writer, sheet_name=excel_sheet_name, index=False)
                
                # Formatação
                workbook = writer.book
                worksheet = writer.sheets[excel_sheet_name]
                
                # Formata cabeçalho
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                
                # Ajusta largura das colunas
                for column in worksheet.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = min(max_length + 2, 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        excel_content = output.getvalue()
        output.close()
        return excel_content
    
    def _export_to_json(self, data: Any) -> bytes:
        """Exporta dados para formato JSON."""
        json_content = json.dumps(data, ensure_ascii=False, indent=2, default=str)
        return json_content.encode('utf-8')
    
    def get_content_type(self, format_type: str) -> str:
        """
        Retorna o content-type apropriado para o formato.
        
        Args:
            format_type: Tipo de formato ('csv', 'xlsx', 'json')
            
        Returns:
            Content-type apropriado
        """
        content_types = {
            'csv': 'text/csv; charset=utf-8',
            'xlsx': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'json': 'application/json; charset=utf-8'
        }
        return content_types.get(format_type, 'application/octet-stream')
    
    def get_file_extension(self, format_type: str) -> str:
        """
        Retorna a extensão de arquivo apropriada.
        
        Args:
            format_type: Tipo de formato ('csv', 'xlsx', 'json')
            
        Returns:
            Extensão de arquivo
        """
        extensions = {
            'csv': '.csv',
            'xlsx': '.xlsx',
            'json': '.json'
        }
        return extensions.get(format_type, '.txt')


# Instância global para uso fácil
data_exporter = DataExporter()